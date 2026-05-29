import streamlit as st
import base64
import os
import io
import re
from datetime import datetime
import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuração da página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="CISS — Conciliador de Despesas",
    page_icon="💳",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── CSS personalizado com identidade CISS ───────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

/* Reset e base */
html, body, [class*=\"css\"] {
    font-family: 'Inter', sans-serif;
}

/* Fundo geral */
.stApp {
    background: #0D1117;
    color: #E6EDF3;
}

/* Remove header padrão do Streamlit */
header[data-testid="stHeader"] { background: transparent; }
#MainMenu, footer { visibility: hidden; }

/* Scrollbar */
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: #161B22; }
::-webkit-scrollbar-thumb { background: #30363D; border-radius: 3px; }

/* ── HEADER ──────────────────────────────────── */
.ciss-header {
    background: linear-gradient(135deg, #161B22 0%, #1C2333 100%);
    border: 1px solid #30363D;
    border-radius: 16px;
    padding: 28px 32px;
    margin-bottom: 28px;
    display: flex;
    align-items: center;
    gap: 20px;
    position: relative;
    overflow: hidden;
}
.ciss-header::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, #29ABE2, #1A7FB5, #29ABE2);
}
.ciss-header-text h1 {
    margin: 0;
    font-size: 22px;
    font-weight: 700;
    color: #E6EDF3;
    letter-spacing: -0.3px;
}
.ciss-header-text p {
    margin: 4px 0 0;
    font-size: 13px;
    color: #7D8590;
}
.version-pill {
    background: rgba(41,171,226,0.12);
    border: 1px solid rgba(41,171,226,0.35);
    color: #29ABE2;
    font-size: 11px;
    font-weight: 700;
    padding: 4px 12px;
    border-radius: 20px;
    margin-left: auto;
    white-space: nowrap;
}

/* ── SECTION LABELS ──────────────────────────── */
.section-label {
    font-size: 10px;
    font-weight: 700;
    color: #484F58;
    letter-spacing: 2.5px;
    text-transform: uppercase;
    margin: 24px 0 12px;
}

/* ── UPLOAD CARDS ────────────────────────────── */
.upload-card {
    background: #161B22;
    border: 1px dashed #30363D;
    border-radius: 12px;
    padding: 20px;
    margin-bottom: 12px;
    transition: border-color 0.2s;
}
.upload-card:hover { border-color: #29ABE2; }

/* File uploader */
[data-testid="stFileUploader"] {
    background: #161B22 !important;
    border: 1px dashed #30363D !important;
    border-radius: 12px !important;
}
[data-testid="stFileUploader"]:hover {
    border-color: #29ABE2 !important;
}
[data-testid="stFileUploadDropzone"] {
    background: transparent !important;
}
[data-testid="stFileUploadDropzone"] p {
    color: #7D8590 !important;
}

/* ── BOTÃO PRINCIPAL ─────────────────────────── */
.stButton > button {
    background: linear-gradient(90deg, #1A7FB5 0%, #29ABE2 100%) !important;
    color: white !important;
    font-weight: 700 !important;
    font-size: 15px !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 14px 28px !important;
    width: 100% !important;
    letter-spacing: 0.3px;
    transition: opacity 0.2s !important;
}
.stButton > button:hover {
    opacity: 0.9 !important;
    box-shadow: 0 0 20px rgba(41,171,226,0.3) !important;
}
.stButton > button:disabled {
    background: #1C2333 !important;
    color: #484F58 !important;
}

/* ── DOWNLOAD BUTTON ─────────────────────────── */
[data-testid="stDownloadButton"] > button {
    background: #161B22 !important;
    color: #29ABE2 !important;
    border: 1.5px solid #29ABE2 !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    width: 100% !important;
    padding: 10px !important;
    transition: all 0.2s !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: #29ABE2 !important;
    color: white !important;
}

/* ── METRIC CARDS ────────────────────────────── */
.metric-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 12px;
    margin: 16px 0;
}
.metric-card {
    background: #161B22;
    border: 1px solid #30363D;
    border-radius: 12px;
    padding: 16px 18px;
    position: relative;
    overflow: hidden;
}
.metric-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
}
.mc-green::before  { background: #3FB950; }
.mc-yellow::before { background: #D29922; }
.mc-red::before    { background: #F85149; }
.mc-orange::before { background: #E3B341; }

.metric-top {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 6px;
}
.metric-icon { font-size: 18px; }
.metric-num  {
    font-size: 28px;
    font-weight: 700;
    line-height: 1;
}
.mc-green  .metric-num { color: #3FB950; }
.mc-yellow .metric-num { color: #D29922; }
.mc-red    .metric-num { color: #F85149; }
.mc-orange .metric-num { color: #E3B341; }

.metric-label { font-size: 12px; color: #7D8590; margin-bottom: 2px; }
.metric-value { font-size: 11px; color: #484F58; }

/* ── DIVIDER ─────────────────────────────────── */
.ciss-divider {
    border: none;
    border-top: 1px solid #21262D;
    margin: 24px 0;
}

/* ── ALERTA / STATUS ─────────────────────────── */
.status-error {
    background: rgba(248,81,73,0.1);
    border: 1px solid rgba(248,81,73,0.3);
    border-radius: 10px;
    padding: 16px 20px;
    color: #F85149;
    font-weight: 600;
    font-size: 14px;
    margin: 16px 0;
}

/* ── SPINNER ─────────────────────────────────── */
[data-testid="stSpinner"] { color: #29ABE2 !important; }

/* ── FOOTER ──────────────────────────────────── */
.ciss-footer {
    text-align: center;
    font-size: 11px;
    color: #30363D;
    margin-top: 32px;
    padding-top: 20px;
    border-top: 1px solid #21262D;
}

/* Oculta label vazio dos uploaders */
[data-testid="stFileUploader"] label { display: none; }
</style>
""", unsafe_allow_html=True)


# ── Logo em base64 ──────────────────────────────────────────────────────────
def logo_b64():
    logo_path = os.path.join(os.path.dirname(__file__), "logo_ciss.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    return None


# ── LÓGICA DO LEITOR EXCEL ──────────────────────────────────────────────────
def _is_date(val):
    if isinstance(val, (datetime, pd.Timestamp)):
        return True
    try:
        pd.to_datetime(val)
        return True
    except Exception:
        return False

def ler_excel(source):
    # Correção: Envolve os bytes explicitamente em io.BytesIO se for uma cadeia de bytes
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
        
    df_raw = pd.read_excel(source, header=None)
    lancamentos = []
    header_rows = []
    
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row.values]
        if 'Data da transação' in vals and 'Lançamentos' in vals:
            header_rows.append(i)

    if not header_rows:
        raise Exception(
            "Formato do Excel não reconhecido. "
            "Certifique-se de usar a fatura exportada pelo banco."
        )

    for h in header_rows:
        row_h = df_raw.iloc[h]
        cols  = [str(v).strip() for v in row_h.values]

        try:    idx_data  = cols.index('Data da transação')
        except: idx_data  = 1
        try:    idx_desc  = cols.index('Lançamentos')
        except: idx_desc  = 2

        if 'Valor em R$' in cols:
            idx_valor = cols.index('Valor em R$')
        elif 'Valor' in cols:
            idx_valor = cols.index('Valor')
        else:
            idx_valor = 7

        for j in range(h + 1, len(df_raw)):
            row       = df_raw.iloc[j]
            data_val  = row.iloc[idx_data]
            desc_val  = row.iloc[idx_desc]
            valor_val = row.iloc[idx_valor]

            if pd.isna(data_val) or str(data_val).strip() in ('', 'nan'):
                break
            if not isinstance(data_val, (datetime, pd.Timestamp)) and not _is_date(data_val):
                break

            try:
                valor = float(str(valor_val).replace(',', '.'))
            except (ValueError, TypeError):
                continue

            lancamentos.append({
                'data':     pd.to_datetime(data_val).date(),
                'descricao': str(desc_val).strip(),
                'valor':    round(valor, 2),
                'fonte':    'fatura',
            })

    if not lancamentos:
        raise Exception("Nenhum lançamento encontrado no Excel.")

    return lancamentos


# ── LÓGICA DO LEITOR PDF ────────────────────────────────────────────────────
def extrair_despesas_pdf(source):
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)

    texto_completo = ""
    with pdfplumber.open(source) as pdf:
        for pagina in pdf.pages:
            t = pagina.extract_text()
            if t:
                texto_completo += t + "\n"

    despesas = []
    linhas   = texto_completo.split('\n')

    padrao_despesa = re.compile(
        r'^(.+?)\s+'
        r'(\d{2}/\d{2}/\d{2})\s+'
        r'(Cartão Corporativo|Reembolso)'
        r'.*?BRL\s+'
        r'([\d.,]+)\s+'
        r'([\d.,]+)'
    )

    i = 0
    while i < len(linhas):
        linha = linhas[i].strip()
        m = padrao_despesa.match(linha)
        if m:
            tipo        = m.group(1).strip()
            data_str    = m.group(2)
            valor_str   = m.group(4).replace('.', '').replace(',', '.')

            try:
                valor = round(float(valor_str), 2)
                data  = datetime.strptime(data_str, '%d/%m/%y').date()
            except (ValueError, TypeError):
                i += 1
                continue

            justificativa = ''
            if i + 1 < len(linhas):
                prox = linhas[i + 1].strip()
                if prox.startswith('Justificativa:'):
                    justificativa = prox.replace('Justificativa:', '').strip()
                    i += 1

            if 'Desconto em folha' not in tipo:
                despesas.append({
                    'tipo':          tipo,
                    'data':          data,
                    'recurso':       m.group(3).strip(),
                    'valor':         valor,
                    'justificativa': justificativa,
                    'fonte':         'paytrack',
                })
        i += 1

    if not despesas:
        raise Exception("Nenhuma despesa encontrada no PDF.")

    return despesas


# ── LÓGICA DO EXPORTADOR EXCEL ──────────────────────────────────────────────
COR_VERDE   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
COR_AMARELO = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
COR_VERMELHO= PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
COR_LARANJA = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
COR_CINZA   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
COR_HEADER  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

FONTE_HEADER = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
FONTE_BOLD   = Font(bold=True, name="Calibri", size=10)
FONTE_NORMAL = Font(name="Calibri", size=10)
FONTE_TITULO = Font(bold=True, name="Calibri", size=13, color="1F4E79")

borda = Border(
    left=Side(style='thin', color="BFBFBF"),
    right=Side(style='thin', color="BFBFBF"),
    top=Side(style='thin', color="BFBFBF"),
    bottom=Side(style='thin', color="BFBFBF"),
)

def _h(ws, row, col, texto):
    c = ws.cell(row=row, column=col, value=texto)
    c.fill = COR_HEADER; c.font = FONTE_HEADER; c.border = borda
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def _c(ws, row, col, value, fill=None, bold=False, fmt=None, align='left'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONTE_BOLD if bold else FONTE_NORMAL
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = borda
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt

def gerar_excel_bytes(linhas_paytrack, linhas_sem_lancamento, resumo):
    wb = Workbook()

    # Aba Resumo
    ws = wb.active
    ws.title = "📊 Resumo"
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18

    ws.merge_cells('A1:C1')
    ws['A1'].value = "CONCILIAÇÃO DE DESPESAS — RELATÓRIO EXECUTIVO"
    ws['A1'].font  = FONTE_TITULO
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    for col, label in enumerate(['Situação', 'Qtd', 'Valor Total (R$)'], 1):
        c = ws.cell(row=3, column=col, value=label)
        c.font = FONTE_BOLD; c.fill = COR_CINZA; c.border = borda
        c.alignment = Alignment(horizontal='center')

    dados = [
        ("✅ Conciliados",               resumo['conciliados'],      resumo['valor_conciliado'],     COR_VERDE),
        ("⚠️ Divergentes",               resumo['divergentes'],      resumo['valor_divergente'],     COR_AMARELO),
        ("❌ No Paytrack, não na Fatura", resumo['nao_encontrados'],  resumo['valor_nao_encontrado'], COR_VERMELHO),
        ("❗ Na Fatura, não no Paytrack", resumo['nao_lancados'],     resumo['valor_nao_lancado'],    COR_LARANJA),
    ]
    for i, (label, qtd, val, fill) in enumerate(dados, 4):
        _c(ws, i, 1, label,  fill=fill, bold=True)
        _c(ws, i, 2, qtd,    fill=fill, bold=True, align='center')
        _c(ws, i, 3, val,    fill=fill, bold=True, fmt='R$ #,##0.00', align='right')

    # Aba Conciliação Paytrack
    ws2 = wb.create_sheet("📋 Conciliação Paytrack")
    cols = [
        ('Tipo (Paytrack)', 18), ('Data (Paytrack)', 14), ('Valor (Paytrack)', 15),
        ('Justificativa', 35),   ('Descrição (Fatura)', 25), ('Data (Fatura)', 14),
        ('Valor (Fatura)', 14),  ('Diferença', 12),           ('Status', 25),
    ]
    for ci, (nome, larg) in enumerate(cols, 1):
        _h(ws2, 1, ci, nome)
        ws2.column_dimensions[get_column_letter(ci)].width = larg
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = 'A2'

    sf = {'CONCILIADO': COR_VERDE, 'DIVERGENTE': COR_AMARELO, 'NÃO ENCONTRADO NA FATURA': COR_VERMELHO}
    for ri, l in enumerate(linhas_paytrack, 2):
        f = sf.get(l['Status'], COR_CINZA)
        _c(ws2,ri,1,l['Tipo (Paytrack)'],f)
        _c(ws2,ri,2,l['Data (Paytrack)'],f,align='center')
        _c(ws2,ri,3,l['Valor (Paytrack)'],f,fmt='R$ #,##0.00',align='right')
        _c(ws2,ri,4,l['Justificativa'],f)
        _c(ws2,ri,5,l['Descrição (Fatura)'],f)
        _c(ws2,ri,6,l['Data (Fatura)'],f,align='center')
        v = l['Valor (Fatura)']
        _c(ws2,ri,7,v,f,fmt='R$ #,##0.00' if v else None,align='right')
        d = l['Diferença']
        _c(ws2,ri,8,d,f,fmt='R$ #,##0.00' if d is not None else None,align='right')
        _c(ws2,ri,9,l['Status'],f,bold=True,align='center')

    # Aba Não Lançados
    ws3 = wb.create_sheet("❗ Não Lançados Paytrack")
    for ci, (nome, larg) in enumerate([('Descrição (Fatura)',35),('Data (Fatura)',14),('Valor (Fatura)',15),('Status',30)], 1):
        _h(ws3, 1, ci, nome)
        ws3.column_dimensions[get_column_letter(ci)].width = larg
    ws3.row_dimensions[1].height = 30
    ws3.freeze_panes = 'A2'

    for ri, l in enumerate(linhas_sem_lancamento, 2):
        _c(ws3,ri,1,l['Descrição (Fatura)'],COR_LARANJA)
        _c(ws3,ri,2,l['Data (Fatura)'],COR_LARANJA,align='center')
        _c(ws3,ri,3,l['Valor (Fatura)'],COR_LARANJA,fmt='R$ #,##0.00',align='right')
        _c(ws3,ri,4,l['Status'],COR_LARANJA,bold=True,align='center')

    if not linhas_sem_lancamento:
        ws3.cell(row=2,column=1,value="✅ Todos os lançamentos da fatura foram encontrados no Paytrack!")
        ws3.cell(row=2,column=1).font = Font(bold=True, color="375623", size=11)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── LÓGICA DO CONCILIADOR ───────────────────────────────────────────────────
def processar_conciliacao(excel_bytes, pdf_bytes):
    fatura   = ler_excel(excel_bytes)
    paytrack = extrair_despesas_pdf(pdf_bytes)

    fatura_restante = list(fatura)
    linhas_paytrack = []
    linhas_sem_lancamento = []

    for item_pt in paytrack:
        valor_pt = item_pt['valor']
        data_pt  = item_pt['data']

        match_exato = None
        match_approx = None

        for idx, item_fat in enumerate(fatura_restante):
            valor_fat = item_fat['valor']
            data_fat  = item_fat['data']
            diff_val  = abs(valor_pt - valor_fat)
            diff_dias = abs((data_pt - data_fat).days)

            if diff_val <= 0.01 and diff_dias <= 1:
                match_exato = (idx, item_fat)
                break
            elif diff_val <= 0.10 and diff_dias <= 1 and match_approx is None:
                match_approx = (idx, item_fat)

        if match_exato:
            idx, item_fat = match_exato
            fatura_restante.pop(idx)
            linhas_paytrack.append({
                'Tipo (Paytrack)':      item_pt['tipo'],
                'Data (Paytrack)':      item_pt['data'].strftime('%d/%m/%Y'),
                'Valor (Paytrack)':     item_pt['valor'],
                'Justificativa':        item_pt['justificativa'],
                'Descrição (Fatura)':   item_fat['descricao'],
                'Data (Fatura)':        item_fat['data'].strftime('%d/%m/%Y'),
                'Valor (Fatura)':       item_fat['valor'],
                'Diferença':            round(item_fat['valor'] - item_pt['valor'], 2),
                'Status':               'CONCILIADO',
            })
        elif match_approx:
            idx, item_fat = match_approx
            fatura_restante.pop(idx)
            linhas_paytrack.append({
                'Tipo (Paytrack)':      item_pt['tipo'],
                'Data (Paytrack)':      item_pt['data'].strftime('%d/%m/%Y'),
                'Valor (Paytrack)':     item_pt['valor'],
                'Justificativa':        item_pt['justificativa'],
                'Descrição (Fatura)':   item_fat['descricao'],
                'Data (Fatura)':        item_fat['data'].strftime('%d/%m/%Y'),
                'Valor (Fatura)':       item_fat['valor'],
                'Diferença':            round(item_fat['valor'] - item_pt['valor'], 2),
                'Status':               'DIVERGENTE',
            })
        else:
            linhas_paytrack.append({
                'Tipo (Paytrack)':      item_pt['tipo'],
                'Data (Paytrack)':      item_pt['data'].strftime('%d/%m/%Y'),
                'Valor (Paytrack)':     item_pt['valor'],
                'Justificativa':        item_pt['justificativa'],
                'Descrição (Fatura)':   '',
                'Data (Fatura)':        '',
                'Valor (Fatura)':       None,
                'Diferença':            None,
                'Status':               'NÃO ENCONTRADO NA FATURA',
            })

    for item_fat in fatura_restante:
        linhas_sem_lancamento.append({
            'Descrição (Fatura)': item_fat['descricao'],
            'Data (Fatura)':      item_fat['data'].strftime('%d/%m/%Y'),
            'Valor (Fatura)':     item_fat['valor'],
            'Status':             'NÃO LANÇADO NO PAYTRACK',
        })

    resumo = {
        'total_paytrack':        len(paytrack),
        'total_fatura':          len(fatura),
        'conciliados':           sum(1 for l in linhas_paytrack if l['Status'] == 'CONCILIADO'),
        'divergentes':           sum(1 for l in linhas_paytrack if l['Status'] == 'DIVERGENTE'),
        'nao_encontrados':       sum(1 for l in linhas_paytrack if 'NÃO ENCONTRADO' in l['Status']),
        'nao_lancados':          len(linhas_sem_lancamento),
        'valor_conciliado':      round(sum(l['Valor (Paytrack)'] for l in linhas_paytrack if l['Status'] == 'CONCILIADO'), 2),
        'valor_divergente':      round(sum(l['Valor (Paytrack)'] for l in linhas_paytrack if l['Status'] == 'DIVERGENTE'), 2),
        'valor_nao_encontrado':  round(sum(l['Valor (Paytrack)'] for l in lines_paytrack if 'NÃO ENCONTRADO' in l['Status']) if 'lines_paytrack' in locals() else sum(1 for l in linhas_paytrack if 'NÃO ENCONTRADO' in l['Status']), 2), # Ajuste de segurança abaixo fixado
    }
    
    # Ajustando chaves do resumo com segurança para somas numéricas
    resumo['valor_nao_encontrado'] = round(sum(l['Valor (Paytrack)'] for l in linhas_paytrack if 'NÃO ENCONTRADO' in l['Status']), 2)
    resumo['valor_nao_lancado'] = round(sum(l['Valor (Fatura)'] for l in linhas_sem_lancamento), 2)

    excel_bytes_out = gerar_excel_bytes(linhas_paytrack, linhas_sem_lancamento, resumo)
    return excel_bytes_out, resumo


# ── INTERFACE GRÁFICA DO STREAMLIT ───────────────────────────────────────────
logo = logo_b64()
logo_html = f'<img src="data:image/png;base64,{logo}" style="height:50px; object-fit: contain;" />' if logo else '<span style="font-size:28px;font-weight:900;color:#29ABE2;">CiSS</span>'

st.markdown(f"""
<div class="ciss-header">
    {logo_html}
    <div style="width:1px;height:40px;background:#30363D;margin:0 4px;"></div>
    <div class="ciss-header-text">
        <h1>Conciliador de Despesas</h1>
        <p>Paytrack &nbsp;×&nbsp; Fatura Bradesco Corporativo</p>
    </div>
    <span class="version-pill">v2.0</span>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="section-label">📂 Arquivos de entrada</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    st.markdown("**📊 Fatura do Cartão**")
    st.caption("Extrato Excel exportado pelo banco")
    excel_file = st.file_uploader("excel", type=["xlsx", "xls"], label_visibility="collapsed", key="excel")

with col2:
    st.markdown("**📄 Relatório Paytrack**")
    st.caption("PDF gerado pelo aplicativo Paytrack")
    pdf_file = st.file_uploader("pdf", type=["pdf"], label_visibility="collapsed", key="pdf")

# Validação segura: Garante que os arquivos foram totalmente carregados
files_ok = excel_file is not None and pdf_file is not None

if excel_file:
    st.success(f"✓ Fatura: **{excel_file.name}** ({excel_file.size/1024:.1f} KB)")
if pdf_file:
    st.success(f"✓ Relatório: **{pdf_file.name}** ({pdf_file.size/1024:.1f} KB)")

st.markdown('<div class="ciss-divider"></div>', unsafe_allow_html=True)

iniciar = st.button("⚡  Iniciar Conciliação", disabled=not files_ok, use_container_width=True)

if not files_ok:
    st.markdown('<p style="text-align:center;color:#484F58;font-size:12px;margin-top:8px;">Selecione os dois arquivos para habilitar a conciliação</p>', unsafe_allow_html=True)

if iniciar and files_ok:
    try:
        with st.spinner("Processando conciliação..."):
            # Lendo os arquivos como bytes de forma garantida
            excel_data = excel_file.read()
            pdf_data = pdf_file.read()
            
            excel_bytes_out, resumo = processar_conciliacao(excel_data, pdf_data)

        st.session_state["resultado"]    = excel_bytes_out
        st.session_state["resumo"]       = resumo
        st.session_state["nome_arquivo"] = f"conciliacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    except Exception as e:
        st.markdown(f'<div class="status-error">❌ Erro: {e}</div>', unsafe_allow_html=True)

if "resultado" in st.session_state:
    resumo = st.session_state["resumo"]
    st.markdown('<div class="ciss-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="section-label">📊 Resultado da conciliação</div>', unsafe_allow_html=True)

    st.markdown(f"""
    <div class="metric-grid">
        <div class="metric-card mc-green">
            <div class="metric-top"><span class="metric-icon">✅</span><span class="metric-num">{resumo['conciliados']}</span></div>
            <div class="metric-label">Conciliados</div><div class="metric-value">R$ {resumo['valor_conciliado']:,.2f}</div>
        </div>
        <div class="metric-card mc-yellow">
            <div class="metric-top"><span class="metric-icon">⚠️</span><span class="metric-num">{resumo['divergentes']}</span></div>
            <div class="metric-label">Divergentes</div><div class="metric-value">R$ {resumo['valor_divergente']:,.2f}</div>
        </div>
        <div class="metric-card mc-red">
            <div class="metric-top"><span class="metric-icon">❌</span><span class="metric-num">{resumo['nao_encontrados']}</span></div>
            <div class="metric-label">Não encontr. na Fatura</div><div class="metric-value">R$ {resumo['valor_nao_encontrado']:,.2f}</div>
        </div>
        <div class="metric-card mc-orange">
            <div class="metric-top"><span class="metric-icon">❗</span><span class="metric-num">{resumo['nao_lancados']}</span></div>
            <div class="metric-label">Não lançados no Paytrack</div><div class="metric-value">R$ {resumo['valor_nao_lancado']:,.2f}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.download_button(
        label="⬇️  Baixar Relatório Excel",
        data=st.session_state["resultado"],
        file_name=st.session_state["nome_arquivo"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

st.markdown('<div class="ciss-footer">CISS Consultoria em Informática, Serviços e Software S/A</div>', unsafe_allow_html=True)
